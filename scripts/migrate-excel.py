#!/usr/bin/env python3
"""
Migra os dados das planilhas v1 (Planilha_CRM_Guilds.xlsx) e v2
(Guilds_Planilha_Controle_Comercial_2026_v2.xlsx) para um arquivo
supabase/seed.sql que pode ser rodado direto no Supabase.

Regras:
- Marcos Oliveira / Fernanda Castro: pegar versão V2 (mais completa)
- Resto do v1 (42 leads): migrar tudo
- 5 leads-exemplo do v2 são marcados is_demo=true
- "Sem Interesse" (v1) → crm_stage='Perdido'
- "Sem Resposta" (v1) → crm_stage='Nutrição'
- "Cadência" (v1) → crm_stage='Prospecção'
- "Raio-X Feito" (v1) → crm_stage='Raio-X Feito'

Como rodar:
    python scripts/migrate-excel.py V1_PATH V2_PATH > supabase/seed.sql
"""
import sys, json
from openpyxl import load_workbook
from datetime import datetime, date

V1 = sys.argv[1] if len(sys.argv) > 1 else 'Planilha_CRM_Guilds.xlsx'
V2 = sys.argv[2] if len(sys.argv) > 2 else 'Guilds_Planilha_Controle_Comercial_2026_v2.xlsx'

def sql_str(v):
    if v is None: return 'null'
    if isinstance(v, bool): return 'true' if v else 'false'
    if isinstance(v, (int, float)): return str(v)
    if isinstance(v, (datetime, date)): return f"'{v.strftime('%Y-%m-%d')}'"
    s = str(v).replace("'", "''").strip()
    return f"'{s}'" if s else 'null'

def map_stage_v1(etapa):
    if not etapa: return ('arquivado', None)
    e = etapa.strip()
    m = {
        'Sem Interesse':   ('arquivado', 'Perdido'),
        'Sem Resposta':    ('arquivado', 'Nutrição'),
        'Cadência':        ('pipeline',  'Prospecção'),
        'Raio-X Feito':    ('pipeline',  'Raio-X Feito'),
    }
    return m.get(e, ('pipeline', 'Prospecção'))

# ===== Lê V1 PIPELINE =====
v1_leads = []
wb1 = load_workbook(V1, data_only=True)
ws = wb1['PIPELINE']
header = None
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i == 2:
        header = list(row)
    elif i >= 3 and any(c not in (None,'') for c in row[:6]):
        d = dict(zip(header, row))
        v1_leads.append(d)

# ===== Lê V2 PIPELINE / RAIO-X / CADENCIA / NEWSLETTER =====
wb2 = load_workbook(V2, data_only=True)

def read_v2(sheet, header_row=4, key_col=2):
    ws = wb2[sheet]
    h = None
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == header_row: h = list(row)
        elif i > header_row and row[key_col] is not None:
            out.append(dict(zip(h, row)))
    return out

v2_pipe   = read_v2('PIPELINE')
v2_raiox  = read_v2('RAIO-X', key_col=1)
v2_cad    = read_v2('CADENCIA', key_col=1)
v2_nl     = read_v2('NEWSLETTER', key_col=1)

# Quem do v1 já está no v2?
v2_names = {l['Nome'].strip().lower(): l for l in v2_pipe if l.get('Nome')}

# ===== Monta seed.sql =====
out = []
out.append("-- ==========================================================")
out.append("-- GUILDS COMERCIAL — SEED DATA")
out.append("-- Gerado por scripts/migrate-excel.py")
out.append("-- 5 leads marcados is_demo (Marcos, Fernanda, Rafaela, Patrícia, João)")
out.append("--   → para apagar:  delete from leads where is_demo = true;")
out.append("-- 42 leads reais migrados do v1 (Planilha_CRM_Guilds.xlsx)")
out.append("-- ==========================================================\n")

out.append("-- IMPORTANTE: rodar APÓS criar os usuários no Supabase Auth")
out.append("-- e popular a tabela profiles com os ids correspondentes.\n")

# Helper: pegar profile_id do responsável (referência por display_name)
out.append("-- Carrega ids dos profiles para usar nas inserções")
out.append("do $$")
out.append("declare")
out.append("  uid_gustavo  uuid := (select id from public.profiles where display_name='Gustavo'   limit 1);")
out.append("  uid_comerc   uuid := (select id from public.profiles where display_name='Comercial' limit 1);")
out.append("  uid_sdr1     uuid := (select id from public.profiles where display_name='SDR 1'     limit 1);")
out.append("  uid_sdr2     uuid := (select id from public.profiles where display_name='SDR 2'     limit 1);")
out.append("begin\n")

# ====== INSERT V2 (5 leads de exemplo) ======
out.append("-- ----- 5 leads-exemplo do v2 (is_demo = true) -----")
for l in v2_pipe:
    cols = []
    vals = []
    def add(c, v): cols.append(c); vals.append(v)
    add('legacy_id',       sql_str(f"v2#{l.get('ID')}"))
    add('is_demo',         'true')
    add('nome',            sql_str(l.get('Nome')))
    add('empresa',         sql_str(l.get('Empresa')))
    add('cargo',           sql_str(l.get('Cargo')))
    add('email',           sql_str(l.get('Email')))
    add('whatsapp',        sql_str(l.get('WhatsApp')))
    add('linkedin',        sql_str(l.get('LinkedIn')))
    add('segmento',        sql_str(l.get('Segmento')))
    add('cidade_uf',       sql_str(l.get('Cidade/UF')))
    add('responsavel_id',  'uid_comerc')
    add('motion',          sql_str(l.get('Motion')))
    add('fonte',           sql_str(l.get('Fonte')))
    add('temperatura',     sql_str(l.get('Temperatura') or 'Frio'))
    add('funnel_stage',    "'pipeline'")
    add('crm_stage',       sql_str(l.get('Etapa CRM')))
    add('decisor',         'true' if (l.get('Decisor?')=='Sim') else 'null')
    add('dor_principal',   sql_str(l.get('Dor principal')))
    add('observacoes',     sql_str(l.get('Observações')))
    add('canal_principal', sql_str(l.get('Canal principal')))
    add('data_primeiro_contato', sql_str(l.get('1º contato')))
    add('data_ultimo_toque',     sql_str(l.get('Último toque')))
    add('proxima_acao',          sql_str(l.get('Próx. ação')))
    add('data_proxima_acao',     sql_str(l.get('Data próx. ação')))
    add('valor_potencial',       sql_str(l.get('Valor potencial (R$)') or 0))
    add('probabilidade',         sql_str(l.get('Probab. %') or 0))
    add('data_proposta',         sql_str(l.get('Data proposta')))
    add('data_fechamento',       sql_str(l.get('Data fechamento')))
    add('newsletter_optin',      'true' if (l.get('Newsletter')=='Sim') else 'false')
    out.append(f"  insert into public.leads ({', '.join(cols)}) values ({', '.join(vals)});")

# ====== INSERT V1 (42 leads reais, exceto Marcos/Fernanda) ======
out.append("\n-- ----- 42 leads reais do v1 -----")
v1_count = 0
for l in v1_leads:
    nome = (l.get('NOME') or '').strip()
    if nome.lower() in v2_names:
        continue  # já migrado via v2
    funnel, crm = map_stage_v1(l.get('ETAPA'))
    cols, vals = [], []
    def add(c, v): cols.append(c); vals.append(v)
    add('legacy_id',       sql_str(f"v1#{nome or l.get('EMPRESA')}"))
    add('is_demo',         'false')
    add('nome',            sql_str(nome) if nome else 'null')
    add('empresa',         sql_str(l.get('EMPRESA')))
    add('cargo',           sql_str(l.get('CARGO')))
    add('whatsapp',        sql_str(l.get('WHATSAPP')))
    add('linkedin',        sql_str(l.get('LINKEDIN')))
    add('segmento',        sql_str(l.get('SETOR')))
    add('responsavel_id',  'uid_comerc')
    add('motion',          "'Outbound'")
    add('fonte',           "'Lista fria'")
    add('temperatura',     "'Frio'")
    add('funnel_stage',    sql_str(funnel))
    add('crm_stage',       sql_str(crm))
    add('observacoes',     sql_str(l.get('OBSERVAÇÕES')))
    pc = l.get('1º CONTATO')
    if isinstance(pc, str):
        try: pc = datetime.strptime(pc, '%d/%m/%Y').date()
        except Exception: pc = None
    add('data_primeiro_contato', sql_str(pc))
    add('data_ultimo_toque',     sql_str(pc))
    pa = l.get('PRÓX. AÇÃO')
    da = l.get('DATA PRÓX. AÇÃO')
    if isinstance(da, str):
        try: da = datetime.strptime(da, '%d/%m/%Y').date()
        except Exception: da = None
    add('proxima_acao',          sql_str(pa))
    add('data_proxima_acao',     sql_str(da))
    out.append(f"  insert into public.leads ({', '.join(cols)}) values ({', '.join(vals)});")
    v1_count += 1

# ====== RAIO-X (do v2) ======
out.append("\n-- ----- Raio-X (do v2 — leads de exemplo) -----")
for r in v2_raiox:
    nivel_raw = (r.get('Nível') or 'Pendente')
    if 'Alto' in nivel_raw:    nivel = 'Alto'
    elif 'Médio' in nivel_raw: nivel = 'Médio'
    elif 'Baixo' in nivel_raw: nivel = 'Baixo'
    else: nivel = 'Pendente'
    nome = r.get('Nome')
    out.append(
        f"  insert into public.raio_x (lead_id, responsavel_id, data_oferta, preco_lista, "
        f"voucher_desconto, gratuito, pago, data_pagamento, score, perda_anual_estimada, "
        f"nivel, saida_recomendada, call_revisao, data_call, observacoes) "
        f"select id, uid_comerc, {sql_str(r.get('Data oferta'))}, "
        f"{sql_str(r.get('Preço lista (R$)') or 97)}, {sql_str(r.get('Voucher desconto (R$)') or 0)}, "
        f"{'true' if r.get('Gratuito?')=='Sim' else 'false'}, "
        f"{'true' if r.get('Pago?')=='Sim' else 'false'}, {sql_str(r.get('Data pagamento'))}, "
        f"{sql_str(r.get('Score'))}, {sql_str(r.get('Perda anual estimada (R$)'))}, "
        f"{sql_str(nivel)}, {sql_str(r.get('Saída recomendada'))}, "
        f"{'true' if r.get('Call revisão?')=='Sim' else 'false'}, {sql_str(r.get('Data call'))}, "
        f"{sql_str(r.get('Observações'))} "
        f"from public.leads where nome = {sql_str(nome)} and is_demo = true limit 1;"
    )

# ====== CADENCIA (do v2) ======
out.append("\n-- ----- Cadência (do v2) -----")
for c in v2_cad:
    nome = c.get('Nome')
    base = c.get('Data base D0')
    canal = c.get('Canal principal')
    for passo in ['D0','D3','D7','D11','D16','D30']:
        obj = c.get(f'{passo} objetivo')
        prev = c.get(f'{passo} previsto')
        st = c.get(f'{passo} status')
        if obj or prev:
            out.append(
                f"  insert into public.cadencia (lead_id, passo, canal, objetivo, data_prevista, status) "
                f"select id, {sql_str(passo)}, {sql_str(canal)}, {sql_str(obj)}, "
                f"{sql_str(prev)}, {sql_str(st or 'pendente')} "
                f"from public.leads where nome = {sql_str(nome)} and is_demo = true limit 1 "
                f"on conflict (lead_id, passo) do nothing;"
            )

# ====== NEWSLETTER (do v2) ======
out.append("\n-- ----- Newsletter (do v2) -----")
for n in v2_nl:
    nome = n.get('Nome')
    out.append(
        f"  insert into public.newsletter (lead_id, responsavel_id, optin, data_entrada, "
        f"ultima_edicao_enviada, proxima_edicao_sugerida, status, cta_provavel, observacoes) "
        f"select id, uid_comerc, {'true' if n.get('Opt-in?')=='Sim' else 'false'}, "
        f"{sql_str(n.get('Data entrada'))}, {sql_str(n.get('Última edição enviada'))}, "
        f"{sql_str(n.get('Próxima edição sugerida'))}, {sql_str(n.get('Status') or 'Ativo')}, "
        f"{sql_str(n.get('CTA mais provável'))}, {sql_str(n.get('Observações'))} "
        f"from public.leads where nome = {sql_str(nome)} and is_demo = true limit 1;"
    )

# ====== METAS (12 semanas + 9 meses do v2) ======
out.append("\n-- ----- Metas semanais (12 semanas) -----")
ws = wb2['METAS']
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i in (1,2,3,4): continue
    if i > 17: break
    if not row[0]: continue
    out.append(
        f"  insert into public.meta_semanal (inicio, fim, meta_leads, meta_resp, meta_raiox, "
        f"meta_calls, meta_props, meta_fech) values ("
        f"{sql_str(row[0])}, {sql_str(row[1])}, {row[2] or 15}, {row[4] or 4}, "
        f"{row[6] or 2}, {row[8] or 2}, {row[10] or 1}, {row[12] or 1}) "
        f"on conflict (inicio) do nothing;"
    )

out.append("\n-- ----- Metas mensais (9 meses) -----")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i < 19: continue
    if i > 30: break
    if not row[0]: continue
    out.append(
        f"  insert into public.meta_mensal (rotulo, inicio, fim, meta_leads, meta_raiox, "
        f"meta_calls, meta_props, meta_fech) values ("
        f"{sql_str(row[0])}, {sql_str(row[1])}, {sql_str(row[2])}, {row[3] or 60}, "
        f"{row[5] or 8}, {row[7] or 6}, {row[9] or 3}, {row[11] or 1}) "
        f"on conflict (inicio) do nothing;"
    )

out.append("\nend $$;\n")
out.append("-- ==========================================================")
out.append(f"-- Total: {len(v2_pipe)} leads-demo + {v1_count} leads reais migrados")
out.append("-- ==========================================================")

print('\n'.join(out))
